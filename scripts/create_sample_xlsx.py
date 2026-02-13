#!/usr/bin/env python3
"""Script to create sample Excel BoM files with edge cases."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path


def create_complex_bom():
    """Create an Excel BoM with merged cells and non-line-1 headers."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bill of Materials"

    # Add some metadata rows before the actual BoM
    ws["A1"] = "Project Name: Demo PCB Board"
    ws["A1"].font = Font(bold=True, size=14)

    ws["A2"] = "Revision: 1.0"
    ws["A3"] = "Date: 2026-02-13"

    # Empty row
    # Row 5 will be blank

    # Headers start at row 6
    headers = ["Ref Des", "Quantity", "Mfr", "Manufacturer Part Number", "Description", "Footprint", "Value"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=6, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Add BoM data
    bom_data = [
        ["R30", 1, "Vishay", "CRCW08051K00FKEA", "Resistor 1k 1%", "0805", "1k"],
        ["R31-R34", 4, "Vishay", "CRCW080510K0FKEA", "Resistor 10k 1%", "0805", "10k"],
        ["C30", 1, "Samsung", "CL10A105KA8NNNC", "MLCC 1uF", "0603", "1uF"],
        ["C31", 1, "Samsung", "CL10A106KQ8NNNC", "MLCC 10uF", "0603", "10uF"],
        ["U30", 1, "Texas Instruments", "LM358DR", "Op-Amp Dual", "SOIC-8", ""],
        ["U31", 1, "Nordic Semi", "NRF52832-QFAA-R", "BLE SoC", "QFN-48", ""],
        ["J30", 1, "Wurth", "629105136821", "USB Type-C Connector", "SMD", ""],
        ["D30", 2, "Diodes Inc", "1N4148W-7-F", "Switching Diode", "SOD-123", "100V"],
    ]

    for row_idx, row_data in enumerate(bom_data, start=7):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save file
    output_path = Path(__file__).parent.parent / "data" / "sample_boms" / "complex_layout.xlsx"
    wb.save(output_path)
    print(f"Created: {output_path}")


def create_merged_cells_bom():
    """Create an Excel BoM with merged cells."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BoM"

    # Title with merged cells
    ws.merge_cells("A1:G1")
    ws["A1"] = "BILL OF MATERIALS - PCB ASSEMBLY"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="center")

    # Headers at row 3
    headers = ["Reference", "Qty", "Manufacturer", "MPN", "Desc", "Package", "Value"]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=3, column=col_idx, value=header)

    # Sample data
    data = [
        ["R40", 1, "Yageo", "RC0805FR-071KL", "1k Resistor", "0805", "1k"],
        ["R41", 1, "Yageo", "RC0805FR-0710KL", "10k Resistor", "0805", "10k"],
        ["C40", 1, "Murata", "GRM188R71C104KA01", "100nF Cap", "0603", "100nF"],
    ]

    for row_idx, row_data in enumerate(data, start=4):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    output_path = Path(__file__).parent.parent / "data" / "sample_boms" / "merged_cells.xlsx"
    wb.save(output_path)
    print(f"Created: {output_path}")


if __name__ == "__main__":
    create_complex_bom()
    create_merged_cells_bom()
    print("All sample Excel files created successfully!")
